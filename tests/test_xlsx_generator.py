from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from app import database
from app.resource_paths import resolve_resource_path
from app.services.template_profile import load_template_profile
from app.services.xlsx_generator import ScoreWorkbookExportError, generate_score_workbook


SECTION_ROWS = {
    "A-1": range(3, 6),
    "A-2": range(6, 11),
    "A-3": range(11, 17),
    "A-4": range(17, 25),
    "A-5": range(25, 31),
    "A-6": range(31, 36),
    "A-7": range(36, 41),
    "A-8": range(41, 44),
}


class XlsxGeneratorTests(unittest.TestCase):
    def test_generates_dynamic_complete_score_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            database_path = root / "data" / "app.db"
            storage_path = root / "storage"
            with patch.dict(
                os.environ,
                {
                    "FULUA_DATABASE_PATH": str(database_path),
                    "FULUA_STORAGE_PATH": str(storage_path),
                },
            ):
                database.init_db()
                project = database.create_project("动态打分表")
                self._seed_complete_project(project["id"])
                output = generate_score_workbook(project["id"])

            self.assertTrue(output.is_file())
            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "整体测评",
                    "说明",
                    "1物理和环境安全",
                    "2网络和通信安全",
                    "3设备和计算安全",
                    "4应用和数据安全",
                    "5管理制度",
                    "6人员管理",
                    "7建设运行",
                    "8应急处置",
                ],
            )
            self.assertEqual(workbook["2网络和通信安全"].max_row, 169)
            a7 = workbook["7建设运行"]
            self.assertEqual(a7.max_row, 9)
            self.assertIn("商用密码应用安全性评估报告", a7["B3"].value)
            self.assertIn("管理体系", a7["B4"].value)
            self.assertEqual(a7["C5"].value, '=IF(COUNTIF(C$3:C$4,"N/A")=COUNTA(C$3:C$4),"N/A",AVERAGE(C$3:C$4))')
            self.assertEqual(workbook["整体测评"]["H6"].value, "='2网络和通信安全'!$K$5")
            self.assertTrue(workbook.calculation.fullCalcOnLoad)
            self.assertTrue(workbook.calculation.forceFullCalc)
            self.assertEqual(workbook.calculation.calcMode, "auto")
            self.assertTrue(all(worksheet.freeze_panes is None for worksheet in workbook.worksheets))
            self.assertGreaterEqual(len(workbook["2网络和通信安全"].data_validations.dataValidation), 3)
            self.assertFalse(any(
                marker in str(cell.value)
                for worksheet in workbook.worksheets
                for row in worksheet.iter_rows()
                for cell in row
                for marker in ("“", "”", "#REF!", "#NAME?")
            ))

    def test_user_object_name_is_written_as_text_and_export_paths_are_unique(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            with patch.dict(
                os.environ,
                {
                    "FULUA_DATABASE_PATH": str(root / "app.db"),
                    "FULUA_STORAGE_PATH": str(root / "storage"),
                },
            ):
                database.init_db()
                project = database.create_project("公式注入防护")
                self._seed_complete_project(project["id"], first_object_name="=HYPERLINK(\"https://example.invalid\",\"x\")")
                first_output = generate_score_workbook(project["id"])
                second_output = generate_score_workbook(project["id"])

            self.assertNotEqual(first_output, second_output)
            workbook = load_workbook(first_output, data_only=False)
            cell = workbook["1物理和环境安全"]["B5"]
            self.assertEqual(cell.value, '=HYPERLINK("https://example.invalid","x")')
            self.assertEqual(cell.data_type, "s")

    def test_incomplete_project_returns_structured_issues(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            with patch.dict(
                os.environ,
                {
                    "FULUA_DATABASE_PATH": str(root / "app.db"),
                    "FULUA_STORAGE_PATH": str(root / "storage"),
                },
            ):
                database.init_db()
                project = database.create_project("未完成")
                with self.assertRaises(ScoreWorkbookExportError) as raised:
                    generate_score_workbook(project["id"])
            self.assertTrue(raised.exception.issues)
            self.assertIn("section_code", raised.exception.issues[0])
            self.assertIn("field", raised.exception.issues[0])

    def test_rejects_empty_unit_and_duplicate_management_object_rows(self) -> None:
        workbook = load_workbook(resolve_resource_path("templates", "scoring", "scoring_template_v1.xlsx"), data_only=False)
        overall = workbook["整体测评"]
        expected_units = [overall.cell(row=row, column=3).value for row in SECTION_ROWS["A-7"]]
        profile = load_template_profile()
        object_names = next(item for item in profile["sections"] if item["code"] == "A-7")["fixed_object_names"]
        rows = [
            {"unit": unit, "object_name": object_name, "compliance": "符合", "unit_score": "1.0000"}
            for unit in expected_units
            for object_name in object_names
        ]
        rows.append({"unit": "", "object_name": object_names[0], "compliance": "符合", "unit_score": "1.0000"})
        rows.append(dict(rows[0]))

        from app.services.xlsx_generator.generator import _validate_section_rows

        issues = _validate_section_rows("A-7", "management", rows, expected_units, profile)
        self.assertTrue(any(issue["field"] == "unit" for issue in issues))
        self.assertTrue(any("重复" in issue["message"] for issue in issues))

    def _seed_complete_project(self, project_id: int, *, first_object_name: str | None = None) -> None:
        workbook = load_workbook(resolve_resource_path("templates", "scoring", "scoring_template_v1.xlsx"), data_only=False)
        overall = workbook["整体测评"]
        profile = load_template_profile()
        for code, overall_rows in SECTION_ROWS.items():
            units = [overall.cell(row=row, column=3).value for row in overall_rows]
            section_profile = next(item for item in profile["sections"] if item["code"] == code)
            rows: list[dict[str, object]] = []
            if code in {"A-1", "A-2", "A-3", "A-4"}:
                object_count = 33 if code == "A-2" else 1
                for unit in units:
                    for object_index in range(object_count):
                        object_name = f"对象{object_index + 1}"
                        if first_object_name is not None and code == "A-1" and unit == units[0] and object_index == 0:
                            object_name = first_object_name
                        rows.append(
                            {
                                "unit": unit,
                                "object_name": object_name,
                                "record_text": "评分记录",
                                "metric_result": {"d": "√", "a": "√", "k": "√", "ra": "1", "rk": "1"},
                            }
                        )
            else:
                for unit in units:
                    for object_name in section_profile["fixed_object_names"]:
                        rows.append(
                            {
                                "unit": unit,
                                "object_name": object_name,
                                "record_text": "评分记录",
                                "metric_result": {"compliance": "符合", "unit_score": "9.9999"},
                            }
                        )
            database.replace_section_rows(project_id, code, rows)


if __name__ == "__main__":
    unittest.main()
