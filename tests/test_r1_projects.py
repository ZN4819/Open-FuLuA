from __future__ import annotations

import hashlib
import asyncio
import json
import os
import re
import sqlite3
import tempfile
import unittest
import uuid
import zipfile
from pathlib import Path
from unittest.mock import patch
from xml.etree import ElementTree as ET

from PIL import Image
from openpyxl import load_workbook

from app import database
from app.config import settings
from app.contracts import (
    FULL_REPORT_TEMPLATE_EDITION,
    FULL_REPORT_TEMPLATE_PACKAGE_ID,
    FULL_REPORT_TEMPLATE_REVISION,
)
from app.services.projects import (
    ProjectServiceError,
    create_typed_project,
    recover_abandoned_upgrade_operations,
    transition_workflow,
    upgrade_project_copy,
)
from app.services.xlsx_generator import generate_score_workbook
from tests.test_xlsx_generator import XlsxGeneratorTests
from app.services.docx_generator import generate_project_docx
from app.services.report_templates.registry import (
    ReportTemplateRegistry,
    ReportTemplateUnavailable,
    report_template_registry,
)


class R1ProjectTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.previous_data_dir = os.environ.get("FULUA_DATA_DIR")
        self.previous_database = os.environ.get("FULUA_DATABASE_PATH")
        os.environ["FULUA_DATA_DIR"] = str(self.root)
        os.environ["FULUA_DATABASE_PATH"] = str(self.root / "data" / "app.db")
        database.init_db()

    def tearDown(self) -> None:
        if self.previous_data_dir is None:
            os.environ.pop("FULUA_DATA_DIR", None)
        else:
            os.environ["FULUA_DATA_DIR"] = self.previous_data_dir
        if self.previous_database is None:
            os.environ.pop("FULUA_DATABASE_PATH", None)
        else:
            os.environ["FULUA_DATABASE_PATH"] = self.previous_database
        self.temporary.cleanup()

    @staticmethod
    def _full_report_arguments() -> dict[str, str]:
        return {
            "template_package_id": FULL_REPORT_TEMPLATE_PACKAGE_ID,
            "template_edition": FULL_REPORT_TEMPLATE_EDITION,
            "template_revision": FULL_REPORT_TEMPLATE_REVISION,
        }

    def test_schema_four_and_typed_creation_preserve_legacy_contract(self) -> None:
        legacy = database.create_project("旧客户端项目")
        full = create_typed_project(
            "完整报告项目",
            project_type="full_report",
            **self._full_report_arguments(),
        )

        self.assertEqual(database.read_schema_version(), "4")
        self.assertEqual(legacy["project_type"], "appendix_a")
        self.assertEqual(legacy["workflow_status"], "draft")
        self.assertIsNone(legacy["template_package_id"])
        uuid.UUID(legacy["project_uuid"])
        self.assertEqual(full["project_type"], "full_report")
        self.assertEqual(full["template_package_id"], FULL_REPORT_TEMPLATE_PACKAGE_ID)
        self.assertEqual(full["template_asset_set_hash"], report_template_registry.load(force=True).asset_set_hash)
        self.assertEqual(len(database.list_sections(legacy["id"])), 8)
        self.assertEqual(len(database.list_sections(full["id"])), 8)

        with self.assertRaises(ProjectServiceError) as error:
            create_typed_project(
                "错误模板",
                project_type="full_report",
                template_package_id="latest",
                template_edition="2023",
                template_revision="2025-12-08",
            )
        self.assertEqual(error.exception.code, "UNSUPPORTED_TEMPLATE_IDENTITY")
        self.assertEqual(len(database.list_projects()), 2)

        with patch.object(
            report_template_registry,
            "load",
            side_effect=ReportTemplateUnavailable("REPORT_TEMPLATE_HASH_MISMATCH", "manifest.json"),
        ):
            with self.assertRaises(ProjectServiceError) as untrusted:
                create_typed_project(
                    "不可信模板",
                    project_type="full_report",
                    **self._full_report_arguments(),
                )
        self.assertEqual(untrusted.exception.code, "TEMPLATE_PACKAGE_UNTRUSTED")
        self.assertEqual(len(database.list_projects()), 2)

        with database.connect() as connection, self.assertRaisesRegex(
            sqlite3.IntegrityError, "PROJECT_IDENTITY_IMMUTABLE"
        ):
            connection.execute(
                "UPDATE projects SET project_uuid = ? WHERE id = ?",
                (str(uuid.uuid4()), int(full["id"])),
            )

    def test_workflow_forward_is_blocked_and_reopen_returns_to_draft(self) -> None:
        project = create_typed_project(
            "工作流项目",
            project_type="full_report",
            **self._full_report_arguments(),
        )
        for action in ("ready-for-review", "confirm"):
            with self.subTest(action=action), self.assertRaises(ProjectServiceError) as error:
                transition_workflow(project["project_uuid"], action)
            self.assertEqual(error.exception.code, "REPORT_VALIDATION_NOT_AVAILABLE")
            self.assertEqual(error.exception.status_code, 409)

        database.update_project_workflow(project["project_uuid"], "ready_for_review")
        reopened = transition_workflow(project["project_uuid"], "reopen")
        self.assertEqual(reopened["workflow_status"], "draft")

    def test_project_request_validation_uses_stable_error_contract(self) -> None:
        from fastapi import Request
        from fastapi.exceptions import RequestValidationError
        from pydantic import ValidationError

        from app.main import stable_project_validation_error
        from app.schemas import ProjectCreate, ProjectUpgradeCopyRequest

        source = database.create_project("错误契约源项目")
        request = Request({
            "type": "http",
            "method": "POST",
            "path": "/api/projects",
            "headers": [],
            "query_string": b"",
            "path_params": {},
            "server": ("127.0.0.1", 8000),
            "client": ("127.0.0.1", 1),
            "scheme": "http",
        })
        try:
            ProjectCreate.model_validate({"name": "非法类型", "project_type": "unknown"})
            self.fail("非法项目类型未被 Pydantic 拒绝")
        except ValidationError as exc:
            invalid_type = asyncio.run(
                stable_project_validation_error(request, RequestValidationError(exc.errors()))
            )

        self.assertEqual(invalid_type.status_code, 422)
        self.assertEqual(
            json.loads(invalid_type.body)["detail"],
            {
                "code": "PROJECT_TYPE_INVALID",
                "message": "项目请求参数无效。",
                "project_uuid": None,
                "field": "project_type",
                "details": {"validation_type": "literal_error"},
            },
        )
        upgrade_request = Request({
            **request.scope,
            "path": f"/api/projects/{source['project_uuid']}/upgrade-copy",
            "path_params": {"project_uuid": source["project_uuid"]},
        })
        try:
            ProjectUpgradeCopyRequest.model_validate({
                "name": "非法幂等键",
                **self._full_report_arguments(),
                "idempotency_key": "not-a-uuid",
            })
            self.fail("非法幂等键未被 Pydantic 拒绝")
        except ValidationError as exc:
            invalid_key = asyncio.run(
                stable_project_validation_error(upgrade_request, RequestValidationError(exc.errors()))
            )
        key_detail = json.loads(invalid_key.body)["detail"]
        self.assertEqual(key_detail["code"], "PROJECT_UUID_INVALID")
        self.assertEqual(key_detail["project_uuid"], source["project_uuid"])
        self.assertEqual(key_detail["field"], "idempotency_key")

    def test_upgrade_copy_preserves_raw_appendix_data_and_remaps_ids(self) -> None:
        source = database.create_project("附录 A 源项目")
        source_section = database.get_section(source["id"], "A-1")
        self.assertIsNotNone(source_section)
        timestamp = database.utc_now()
        source_dir = settings.storage_path / "uploads" / str(source["id"]) / "A-1"
        source_dir.mkdir(parents=True)
        source_file = source_dir / "evidence.png"
        Image.new("RGB", (8, 8), color=(34, 84, 124)).save(source_file)
        relative_source = source_file.relative_to(settings.storage_path).as_posix()

        with database.connect() as connection:
            connection.execute(
                """
                INSERT INTO section_subsystems (
                    project_id, section_code, name, sort_order, created_at, updated_at
                ) VALUES (?, 'A-1', '机房子系统', 1, ?, ?)
                """,
                (source["id"], timestamp, timestamp),
            )
            image = connection.execute(
                """
                INSERT INTO evidence_images (
                    project_id, section_code, file_path, original_name, caption, alt_text,
                    sort_order, created_at, updated_at
                ) VALUES (?, 'A-1', ?, 'evidence.png', '证据图', '替代文本', 1, ?, ?)
                """,
                (source["id"], relative_source, timestamp, timestamp),
            )
            source_image_id = int(image.lastrowid)
            row = connection.execute(
                """
                INSERT INTO assessment_rows (
                    section_id, unit, object_name, subsystem, record_text,
                    sort_order, created_at, updated_at
                ) VALUES (?, '身份鉴别', '机房', '机房子系统', ?, 1, ?, ?)
                """,
                (
                    source_section["id"],
                    f"记录 [[FIG:{source_image_id}]]",
                    timestamp,
                    timestamp,
                ),
            )
            source_row_id = int(row.lastrowid)
            connection.execute(
                """
                INSERT INTO metric_results (
                    row_id, d, a, k, ra, rk, object_score, unit_score, compliance
                ) VALUES (?, '√', '×', '/', '0.5', '1.2', '0.5000', '0.5000', '部分符合')
                """,
                (source_row_id,),
            )
            connection.execute(
                """
                INSERT INTO cross_references (
                    source_row_id, target_image_id, token, display_text
                ) VALUES (?, ?, ?, '证据图')
                """,
                (source_row_id, source_image_id, f"[[FIG:{source_image_id}]]"),
            )

        source_updated_at = source["updated_at"]
        source_hash = hashlib.sha256(source_file.read_bytes()).hexdigest()
        idempotency_key = str(uuid.uuid4())
        target = upgrade_project_copy(
            source["project_uuid"],
            name="完整报告副本",
            idempotency_key=idempotency_key,
            **self._full_report_arguments(),
        )
        repeated = upgrade_project_copy(
            source["project_uuid"],
            name="重试名称不会生效",
            idempotency_key=idempotency_key,
            **self._full_report_arguments(),
        )

        self.assertEqual(repeated["id"], target["id"])
        self.assertEqual(target["source_project_uuid"], source["project_uuid"])
        self.assertEqual(target["created_by_operation"], "upgrade_copy")
        with database.connect() as connection:
            target_row = connection.execute(
                """
                SELECT r.*, m.d, m.a, m.k, m.ra, m.rk, m.object_score,
                       m.unit_score, m.compliance
                FROM assessment_rows r
                JOIN appendix_sections s ON s.id = r.section_id
                JOIN metric_results m ON m.row_id = r.id
                WHERE s.project_id = ?
                """,
                (target["id"],),
            ).fetchone()
            target_image = connection.execute(
                "SELECT * FROM evidence_images WHERE project_id = ?",
                (target["id"],),
            ).fetchone()
            target_reference = connection.execute(
                """
                SELECT c.* FROM cross_references c
                JOIN assessment_rows r ON r.id = c.source_row_id
                JOIN appendix_sections s ON s.id = r.section_id
                WHERE s.project_id = ?
                """,
                (target["id"],),
            ).fetchone()
            target_subsystem = connection.execute(
                "SELECT name FROM section_subsystems WHERE project_id = ?",
                (target["id"],),
            ).fetchone()

        self.assertNotEqual(target_row["id"], source_row_id)
        self.assertNotEqual(target_image["id"], source_image_id)
        self.assertEqual(
            (target_row["d"], target_row["a"], target_row["k"], target_row["ra"], target_row["rk"]),
            ("√", "×", "/", "0.5", "1.2"),
        )
        self.assertEqual(target_row["object_score"], "0.5000")
        self.assertEqual(target_row["unit_score"], "0.5000")
        self.assertEqual(target_row["compliance"], "部分符合")
        expected_token = f"[[FIG:{target_image['id']}]]"
        self.assertIn(expected_token, target_row["record_text"])
        self.assertEqual(target_reference["token"], expected_token)
        self.assertEqual(target_reference["target_image_id"], target_image["id"])
        self.assertEqual(target_subsystem["name"], "机房子系统")

        target_file = settings.storage_path / target_image["file_path"]
        self.assertNotEqual(target_file.resolve(), source_file.resolve())
        self.assertEqual(hashlib.sha256(target_file.read_bytes()).hexdigest(), source_hash)
        self.assertEqual(database.get_project_by_id(source["id"])["updated_at"], source_updated_at)
        self.assertEqual(hashlib.sha256(source_file.read_bytes()).hexdigest(), source_hash)

        for mode in ("editable", "final"):
            for project_id in (source["id"], target["id"]):
                exported = generate_project_docx(project_id, mode)
                leaked_factor_tokens: list[str] = []
                visible_text: list[str] = []
                with zipfile.ZipFile(exported) as package:
                    for name in package.namelist():
                        if not name.startswith("word/") or not name.endswith(".xml"):
                            continue
                        root = ET.fromstring(package.read(name))
                        for element in root.iter():
                            local_name = element.tag.rsplit("}", 1)[-1]
                            if local_name == "t" and element.text:
                                visible_text.append(element.text)
                            if local_name not in {"t", "instrText", "tag", "alias"}:
                                continue
                            candidates = [element.text or "", *element.attrib.values()]
                            leaked_factor_tokens.extend(
                                candidate
                                for candidate in candidates
                                if re.search(r"(?<![A-Za-z])R[ak](?![A-Za-z])", candidate)
                            )
                self.assertEqual(leaked_factor_tokens, [], (mode, project_id))
                rendered = "".join(visible_text)
                self.assertIn("机房", rendered)
                self.assertIn("记录", rendered)

        database.delete_project(target["id"])
        from app.services.projects import remove_project_runtime_files

        remove_project_runtime_files(target["id"], target["project_uuid"])
        self.assertTrue(source_file.is_file())
        with self.assertRaises(ProjectServiceError) as gone:
            upgrade_project_copy(
                source["project_uuid"],
                name="不能重建",
                idempotency_key=idempotency_key,
                **self._full_report_arguments(),
            )
        self.assertEqual(gone.exception.code, "IDEMPOTENT_RESULT_GONE")

    def test_upgrade_failure_removes_target_and_can_retry_same_key(self) -> None:
        source = database.create_project("缺图项目")
        timestamp = database.utc_now()
        with database.connect() as connection:
            connection.execute(
                """
                INSERT INTO evidence_images (
                    project_id, section_code, file_path, original_name, caption, alt_text,
                    sort_order, created_at, updated_at
                ) VALUES (?, 'A-1', 'uploads/missing.png', 'missing.png', '', '', 1, ?, ?)
                """,
                (source["id"], timestamp, timestamp),
            )
        key = str(uuid.uuid4())
        with self.assertRaises(ProjectServiceError) as error:
            upgrade_project_copy(
                source["project_uuid"],
                name="失败副本",
                idempotency_key=key,
                **self._full_report_arguments(),
            )
        self.assertEqual(error.exception.code, "SOURCE_EVIDENCE_MISSING")
        self.assertEqual(len(database.list_projects()), 1)
        with database.connect() as connection:
            operation = connection.execute(
                "SELECT status, target_project_id FROM project_upgrade_operations"
            ).fetchone()
        self.assertEqual(operation["status"], "failed")
        self.assertIsNone(operation["target_project_id"])

        missing = settings.storage_path / "uploads" / "missing.png"
        missing.parent.mkdir(parents=True, exist_ok=True)
        missing.write_bytes(b"restored")
        retried = upgrade_project_copy(
            source["project_uuid"],
            name="成功副本",
            idempotency_key=key,
            **self._full_report_arguments(),
        )
        self.assertEqual(retried["project_type"], "full_report")

    def test_abandoned_upgrade_is_recovered_and_same_key_can_retry(self) -> None:
        source = database.create_project("崩溃恢复源项目")
        key = str(uuid.uuid4())
        from app.services import projects as project_service

        target_uuid = project_service._upgrade_target_uuid(source["project_uuid"], key)
        orphan = settings.storage_path / "uploads" / target_uuid / "orphan.txt"
        orphan.parent.mkdir(parents=True)
        orphan.write_text("abandoned", encoding="utf-8")
        staging = project_service._project_upgrade_operation_dir(source["project_uuid"], key)
        staging.mkdir(parents=True)
        (staging / "stale.tmp").write_text("stale", encoding="utf-8")
        with database.connect() as connection:
            connection.execute(
                """
                INSERT INTO project_upgrade_operations (
                    source_project_uuid, idempotency_key, target_project_id,
                    status, lease_id, error_code, created_at, updated_at
                ) VALUES (?, ?, NULL, 'pending', ?, NULL, ?, ?)
                """,
                (
                    source["project_uuid"],
                    key,
                    str(uuid.uuid4()),
                    "2000-01-01T00:00:00+00:00",
                    "2000-01-01T00:00:00+00:00",
                ),
            )

        target = upgrade_project_copy(
            source["project_uuid"],
            name="恢复后的完整报告",
            idempotency_key=key,
            **self._full_report_arguments(),
        )

        self.assertEqual(target["project_uuid"], target_uuid)
        self.assertFalse(orphan.exists())
        self.assertFalse((staging / "stale.tmp").exists())
        with database.connect() as connection:
            operation = connection.execute(
                "SELECT status, lease_id, target_project_id FROM project_upgrade_operations"
            ).fetchone()
        self.assertEqual(operation["status"], "completed")
        self.assertIsNone(operation["lease_id"])
        self.assertEqual(operation["target_project_id"], target["id"])

    def test_startup_recovery_marks_fresh_pending_operation_retryable(self) -> None:
        source = database.create_project("启动恢复源项目")
        key = str(uuid.uuid4())
        with database.connect() as connection:
            connection.execute(
                """
                INSERT INTO project_upgrade_operations (
                    source_project_uuid, idempotency_key, target_project_id,
                    status, lease_id, error_code, created_at, updated_at
                ) VALUES (?, ?, NULL, 'pending', ?, NULL, ?, ?)
                """,
                (
                    source["project_uuid"],
                    key,
                    str(uuid.uuid4()),
                    database.utc_now(),
                    database.utc_now(),
                ),
            )

        self.assertEqual(recover_abandoned_upgrade_operations(), 1)
        target = upgrade_project_copy(
            source["project_uuid"],
            name="启动恢复后的完整报告",
            idempotency_key=key,
            **self._full_report_arguments(),
        )
        self.assertEqual(target["project_type"], "full_report")

    def test_upgrade_copy_preserves_complete_xlsx_semantics(self) -> None:
        source = database.create_project("完整打分表源项目")
        XlsxGeneratorTests()._seed_complete_project(source["id"])
        target = upgrade_project_copy(
            source["project_uuid"],
            name="完整打分表副本",
            idempotency_key=str(uuid.uuid4()),
            **self._full_report_arguments(),
        )

        source_book = load_workbook(generate_score_workbook(source["id"]), data_only=False)
        target_book = load_workbook(generate_score_workbook(target["id"]), data_only=False)
        self.assertEqual(source_book.sheetnames, target_book.sheetnames)
        for sheet_name in source_book.sheetnames:
            source_sheet = source_book[sheet_name]
            target_sheet = target_book[sheet_name]
            self.assertEqual((source_sheet.max_row, source_sheet.max_column), (target_sheet.max_row, target_sheet.max_column))
            self.assertEqual(
                [[cell.value for cell in row] for row in source_sheet.iter_rows()],
                [[cell.value for cell in row] for row in target_sheet.iter_rows()],
                sheet_name,
            )


class R1TemplateRegistryTests(unittest.TestCase):
    def test_clean_checkout_rule_hints_line_endings_are_trusted(self) -> None:
        package = ReportTemplateRegistry().load(force=True)
        self.assertEqual(package.package_id, FULL_REPORT_TEMPLATE_PACKAGE_ID)

    def test_rule_hints_content_tampering_is_still_rejected(self) -> None:
        registry = ReportTemplateRegistry()
        original = registry._root()
        with tempfile.TemporaryDirectory() as temporary_directory:
            copied = Path(temporary_directory) / "package"
            copied.mkdir()
            for name in (
                "asset_hashes.json",
                "runtime_template.docx",
                "field_dictionary.json",
                "manifest.json",
                "rule_hints.json",
                "narrative_templates.json",
            ):
                (copied / name).write_bytes((original / name).read_bytes())
            rule_hints = copied / "rule_hints.json"
            rule_hints.write_bytes(rule_hints.read_bytes().replace(b'"hint_001"', b'"hint_999"', 1))
            with patch.object(registry, "_root", return_value=copied):
                with self.assertRaises(ReportTemplateUnavailable) as error:
                    registry.load(force=True)
        self.assertEqual(error.exception.code, "REPORT_TEMPLATE_HASH_MISMATCH")
        self.assertEqual(error.exception.asset, "rule_hints.json")


if __name__ == "__main__":
    unittest.main()
