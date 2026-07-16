from __future__ import annotations

import re
from collections import Counter, defaultdict
from copy import copy
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList

from ... import database
from ...config import settings
from ...resource_paths import resolve_resource_path
from ..scoring import MANAGEMENT_COMPLIANCE_SCORES
from ..template_profile import load_template_profile


SECTION_LAYOUT = (
    ("A-1", "1物理和环境安全", range(3, 6), "technical"),
    ("A-2", "2网络和通信安全", range(6, 11), "technical"),
    ("A-3", "3设备和计算安全", range(11, 17), "technical"),
    ("A-4", "4应用和数据安全", range(17, 25), "technical"),
    ("A-5", "5管理制度", range(25, 31), "management"),
    ("A-6", "6人员管理", range(31, 36), "management"),
    ("A-7", "7建设运行", range(36, 41), "management"),
    ("A-8", "8应急处置", range(41, 44), "management"),
)
TECHNICAL_VALUES = {"√", "×", "/"}
RA_VALUES = {"1", "0.5", "0.2"}
RK_VALUES = {"1", "1.2"}


class ScoreWorkbookExportError(RuntimeError):
    def __init__(
        self,
        message: str,
        issues: list[dict[str, str]] | None = None,
        *,
        status_code: int = 400,
    ) -> None:
        super().__init__(message)
        self.issues = issues or []
        self.status_code = status_code


def generate_score_workbook(project_id: int) -> Path:
    project = database.get_project_by_id(project_id)
    if project is None:
        raise ScoreWorkbookExportError("项目不存在。", status_code=404)

    template_path = resolve_resource_path("templates", "scoring", "scoring_template_v1.xlsx")
    if not template_path.is_file():
        raise ScoreWorkbookExportError("打分表模板不存在。")

    workbook = load_workbook(template_path)
    expected_units = _expected_units(workbook)
    sections = {section["code"]: section for section in database.list_sections(project_id)}
    section_rows: dict[str, list[dict[str, Any]]] = {}
    issues: list[dict[str, str]] = []
    profile = load_template_profile()

    for code, _, _, table_type in SECTION_LAYOUT:
        section = sections.get(code)
        if section is None:
            issues.append(_issue(code, "", "", "section", "缺少评分章节。"))
            continue
        rows = [dict(row) for row in database.list_effective_assessment_rows(section["id"])]
        section_rows[code] = rows
        issues.extend(_validate_section_rows(code, table_type, rows, expected_units[code], profile))

    if issues:
        raise ScoreWorkbookExportError("项目评分数据未完成，无法导出正式打分表。", issues)

    unit_links: dict[tuple[str, str], tuple[str, str, str]] = {}
    for code, sheet_name, overall_rows, table_type in SECTION_LAYOUT:
        worksheet = workbook[sheet_name]
        rows = section_rows[code]
        units = expected_units[code]
        if table_type == "technical":
            links = _build_technical_sheet(worksheet, code, rows, units, list(overall_rows))
        else:
            section_profile = next(item for item in profile["sections"] if item["code"] == code)
            links = _build_management_sheet(
                worksheet,
                code,
                rows,
                units,
                list(section_profile["fixed_object_names"]),
                list(overall_rows),
            )
        for unit, cells in links.items():
            unit_links[(code, unit)] = (sheet_name, *cells)

    _rebuild_overall_links(workbook["整体测评"], expected_units, unit_links)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = None
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output_path = _export_path(project)
    workbook.save(output_path)
    _validate_generated_workbook(output_path)
    return output_path


def validate_score_workbook_rows(
    section_rows: dict[str, list[dict[str, Any]]],
) -> list[dict[str, str]]:
    """按正式打分表口径校验一组尚未入库的附录 A 行。"""

    expected_units, profile = _score_workbook_validation_context()
    issues: list[dict[str, str]] = []
    for code, _, _, table_type in SECTION_LAYOUT:
        rows = section_rows.get(code)
        if rows is None:
            issues.append(_issue(code, "", "", "section", "缺少评分章节。"))
            continue
        issues.extend(
            _validate_section_rows(
                code,
                table_type,
                rows,
                expected_units[code],
                profile,
            )
        )
    return issues


@lru_cache(maxsize=1)
def _score_workbook_validation_context() -> tuple[dict[str, list[str]], dict[str, Any]]:
    template_path = resolve_resource_path(
        "templates", "scoring", "scoring_template_v1.xlsx"
    )
    workbook = load_workbook(template_path, read_only=True, data_only=False)
    try:
        expected_units = _expected_units(workbook)
    finally:
        workbook.close()
    profile = load_template_profile()
    return expected_units, profile


def _expected_units(workbook: Any) -> dict[str, list[str]]:
    overall = workbook["整体测评"]
    return {
        code: [str(overall.cell(row=row, column=3).value or "").strip() for row in overall_rows]
        for code, _, overall_rows, _ in SECTION_LAYOUT
    }


def _validate_section_rows(
    code: str,
    table_type: str,
    rows: list[dict[str, Any]],
    expected_units: list[str],
    profile: dict[str, Any],
) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row.get("unit") or "").strip()].append(row)

    actual_units = [unit for unit in grouped if unit]
    if "" in grouped:
        issues.append(_issue(code, "", "", "unit", "存在未填写测评指标的评分记录。"))
    if actual_units != expected_units:
        missing = [unit for unit in expected_units if unit not in grouped]
        extra = [unit for unit in actual_units if unit not in expected_units]
        if missing:
            issues.append(_issue(code, "、".join(missing), "", "unit", "缺少模板要求的测评指标。"))
        if extra:
            issues.append(_issue(code, "、".join(extra), "", "unit", "存在模板未定义的测评指标。"))

    fixed_names: Counter[str] = Counter()
    if table_type == "management":
        section_profile = next(item for item in profile["sections"] if item["code"] == code)
        fixed_names = Counter(section_profile["fixed_object_names"])

    for unit in expected_units:
        unit_rows = grouped.get(unit, [])
        if not unit_rows:
            continue
        if table_type == "management":
            actual_names = Counter(str(row.get("object_name") or "").strip() for row in unit_rows)
            if actual_names != fixed_names:
                issues.append(_issue(code, unit, "", "object_name", "固定测评对象存在缺失、重复或不匹配。"))
        for row in unit_rows:
            object_name = str(row.get("object_name") or "").strip()
            if not object_name:
                issues.append(_issue(code, unit, "", "object_name", "缺少测评对象。"))
            if table_type == "technical":
                for field, allowed in (("d", TECHNICAL_VALUES), ("a", TECHNICAL_VALUES), ("k", TECHNICAL_VALUES)):
                    value = str(row.get(field) or "").strip()
                    if value not in allowed:
                        issues.append(_issue(code, unit, object_name, field, f"{field.upper()} 未填写或值不合法。"))
                for field, allowed in (("ra", RA_VALUES), ("rk", RK_VALUES)):
                    value = str(row.get(field) or "").strip()
                    if value not in allowed:
                        issues.append(_issue(code, unit, object_name, field, f"{field.capitalize()} 未填写或值不合法。"))
                if not str(row.get("object_score") or "").strip() or not str(row.get("unit_score") or "").strip():
                    issues.append(_issue(code, unit, object_name, "score", "对象评分或单元得分尚未完成。"))
            else:
                compliance = str(row.get("compliance") or "").strip()
                if compliance not in MANAGEMENT_COMPLIANCE_SCORES:
                    issues.append(_issue(code, unit, object_name, "compliance", "符合情况未填写或值不合法。"))
                if not str(row.get("unit_score") or "").strip():
                    issues.append(_issue(code, unit, object_name, "unit_score", "单元得分尚未完成。"))
    return issues


def _build_technical_sheet(
    worksheet: Any,
    code: str,
    rows: list[dict[str, Any]],
    units: list[str],
    overall_rows: list[int],
) -> dict[str, tuple[str, str]]:
    data_style = [_snapshot_cell(worksheet.cell(row=5, column=column)) for column in range(1, 17)]
    data_height = worksheet.row_dimensions[5].height
    _remove_merges_from_row(worksheet, 5)
    if worksheet.max_row >= 5:
        worksheet.delete_rows(5, worksheet.max_row - 4)
    worksheet.data_validations = DataValidationList()
    worksheet.conditional_formatting = ConditionalFormattingList()

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["unit"]).strip()].append(row)

    current_row = 5
    links: dict[str, tuple[str, str]] = {}
    for unit, overall_row in zip(units, overall_rows, strict=True):
        unit_rows = grouped[unit]
        start = current_row
        end = start + len(unit_rows) - 1
        for offset, source in enumerate(unit_rows):
            row_number = start + offset
            _apply_row_style(worksheet, row_number, data_style, data_height)
            worksheet.cell(row_number, 1, unit if offset == 0 else None)
            _set_text(worksheet.cell(row_number, 2), source["object_name"])
            worksheet.cell(row_number, 3, f'=IF(AND(D{row_number}="/",E{row_number}="/",F{row_number}="/"),"不适用","适用")')
            worksheet.cell(row_number, 4, str(source["d"]))
            worksheet.cell(row_number, 5, str(source["a"]))
            worksheet.cell(row_number, 6, str(source["k"]))
            worksheet.cell(row_number, 7, float(str(source["ra"])))
            worksheet.cell(row_number, 8, float(str(source["rk"])))
            worksheet.cell(row_number, 9, _technical_score_formula(row_number))
            worksheet.cell(row_number, 9).number_format = "0.0000"
            worksheet.cell(row_number, 10, f'=IF(I{row_number}="N/A","不适用",IF(I{row_number}=1,"符合",IF(I{row_number}=0,"不符合","部分符合")))')
            worksheet.cell(row_number, 14, f'=IF(I{row_number}="N/A","N/A",\'整体测评\'!$O${overall_row}/$M${start})')
            worksheet.cell(row_number, 15, f'=IF(I{row_number}="N/A","N/A",I{row_number}*N{row_number})')
            worksheet.cell(row_number, 16, f'=IF(I{row_number}="N/A","N/A",N{row_number}-O{row_number})')
            for column in (14, 15, 16):
                worksheet.cell(row_number, column).number_format = "0.0000"
        worksheet.cell(start, 11, f'=IF(COUNTIF($I${start}:$I${end},"N/A")=COUNTA($I${start}:$I${end}),"N/A",AVERAGE($I${start}:$I${end}))')
        worksheet.cell(start, 11).number_format = "0.0000"
        worksheet.cell(start, 12, f'=IF(K{start}="N/A","不适用",IF(K{start}=1,"符合",IF(K{start}=0,"不符合","部分符合")))')
        worksheet.cell(start, 13, f'=COUNTIF($I${start}:$I${end},"<>N/A")')
        for column in (1, 11, 12, 13):
            if end > start:
                worksheet.merge_cells(start_row=start, start_column=column, end_row=end, end_column=column)
        links[unit] = (f"K{start}", f"L{start}")
        current_row = end + 1

    _add_list_validation(worksheet, f"D5:F{current_row - 1}", ["√", "×", "/"])
    _add_list_validation(worksheet, f"G5:G{current_row - 1}", ["1", "0.5", "0.2"])
    _add_list_validation(worksheet, f"H5:H{current_row - 1}", ["1", "1.2"])
    worksheet.print_area = f"A1:P{current_row - 1}"
    for column in ("N", "O", "P"):
        worksheet.column_dimensions[column].width = max(worksheet.column_dimensions[column].width or 0, 12)
    return links


def _build_management_sheet(
    worksheet: Any,
    code: str,
    rows: list[dict[str, Any]],
    units: list[str],
    object_names: list[str],
    overall_rows: list[int],
) -> dict[str, tuple[str, str]]:
    object_style = [_snapshot_cell(worksheet.cell(row=3, column=column)) for column in range(1, worksheet.max_column + 1)]
    object_height = worksheet.row_dimensions[3].height
    summary_styles = {
        offset: [_snapshot_cell(worksheet.cell(row=offset, column=column)) for column in range(1, worksheet.max_column + 1)]
        for offset in range(4, 9)
    }
    summary_heights = {offset: worksheet.row_dimensions[offset].height for offset in range(4, 9)}
    _remove_merges_from_row(worksheet, 3)
    worksheet.delete_rows(3, worksheet.max_row - 2)
    worksheet.data_validations = DataValidationList()
    worksheet.conditional_formatting = ConditionalFormattingList()

    by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (str(row["unit"]).strip(), str(row["object_name"]).strip())
        if key in by_key:
            raise ScoreWorkbookExportError(
                "管理评分数据包含重复对象，无法生成正式打分表。",
                [_issue(code, key[0], key[1], "object_name", "同一指标下的固定测评对象重复。")],
            )
        by_key[key] = row
    first_object_row = 3
    last_object_row = first_object_row + len(object_names) - 1
    for offset, object_name in enumerate(object_names):
        row_number = first_object_row + offset
        _apply_row_style(worksheet, row_number, object_style, object_height)
        worksheet.cell(row_number, 1, offset + 1)
        worksheet.cell(row_number, 2, object_name)
        for column, unit in enumerate(units, start=3):
            source = by_key[(unit, object_name)]
            score = MANAGEMENT_COMPLIANCE_SCORES[str(source["compliance"]).strip()]
            worksheet.cell(row_number, column, "N/A" if score == "/" else float(score))

    summary_start = last_object_row + 1
    labels = ("测评单元得分Si,j", "单元测评结果（符合/部分符合/不符合/不适用）", "所占分值", "已得分值", "丢失分值")
    for index, label in enumerate(labels):
        row_number = summary_start + index
        source_row = 4 + index
        _apply_row_style(worksheet, row_number, summary_styles[source_row], summary_heights[source_row])
        worksheet.cell(row_number, 1, label)
        worksheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=2)

    links: dict[str, tuple[str, str]] = {}
    for column, (unit, overall_row) in enumerate(zip(units, overall_rows, strict=True), start=3):
        letter = get_column_letter(column)
        score_row = summary_start
        result_row = summary_start + 1
        worksheet.cell(score_row, column, f'=IF(COUNTIF({letter}${first_object_row}:{letter}${last_object_row},"N/A")=COUNTA({letter}${first_object_row}:{letter}${last_object_row}),"N/A",AVERAGE({letter}${first_object_row}:{letter}${last_object_row}))')
        worksheet.cell(score_row, column).number_format = "0.0000"
        worksheet.cell(result_row, column, f'=IF({letter}{score_row}="N/A","不适用",IF({letter}{score_row}=1,"符合",IF({letter}{score_row}=0,"不符合","部分符合")))')
        worksheet.cell(summary_start + 2, column, f'=IF({letter}{result_row}="不适用","N/A",\'整体测评\'!$O${overall_row})')
        worksheet.cell(summary_start + 3, column, f'=IF({letter}{result_row}="不适用","N/A",{letter}{score_row}*{letter}{summary_start + 2})')
        worksheet.cell(summary_start + 4, column, f'=IF({letter}{result_row}="不适用","N/A",{letter}{summary_start + 2}-{letter}{summary_start + 3})')
        for row_number in range(summary_start + 2, summary_start + 5):
            worksheet.cell(row_number, column).number_format = "0.0000"
        links[unit] = (f"{letter}{score_row}", f"{letter}{result_row}")

    last_column = get_column_letter(len(units) + 2)
    _add_list_validation(worksheet, f"C3:{last_column}{last_object_row}", ["1", "0.5", "0", "N/A"])
    worksheet.print_area = f"A1:{last_column}{summary_start + 4}"
    return links


def _rebuild_overall_links(
    worksheet: Any,
    expected_units: dict[str, list[str]],
    unit_links: dict[tuple[str, str], tuple[str, str, str]],
) -> None:
    for code, _, overall_rows, _ in SECTION_LAYOUT:
        for unit, row_number in zip(expected_units[code], overall_rows, strict=True):
            sheet_name, score_cell, result_cell = unit_links[(code, unit)]
            result_ref = f"'{sheet_name}'!${result_cell[0]}${result_cell[1:]}"
            score_ref = f"'{sheet_name}'!${score_cell[0]}${score_cell[1:]}"
            for column in range(4, 8):
                label_cell = worksheet.cell(row=2, column=column).coordinate
                worksheet.cell(row_number, column, f'=IF({result_ref}={label_cell},"√","")')
            worksheet.cell(row_number, 8, f"={score_ref}")
            worksheet.cell(row_number, 8).number_format = "0.0000"
    for column in ("H", "J", "L", "O", "P", "Q"):
        for row_number in range(3, 45):
            worksheet[f"{column}{row_number}"].number_format = "0.0000"
    for column in ("O", "P", "Q"):
        worksheet.column_dimensions[column].width = max(worksheet.column_dimensions[column].width or 0, 13)


def _technical_score_formula(row_number: int) -> str:
    return (
        f'=IF(C{row_number}="不适用","N/A",IF(D{row_number}<>"√",0,'
        f'IF(AND(E{row_number}="√",F{row_number}="√"),1,'
        f'IF(AND(E{row_number}<>"√",F{row_number}="√"),0.5*G{row_number},'
        f'IF(AND(E{row_number}="√",F{row_number}<>"√"),0.5*H{row_number},'
        f'0.25*G{row_number}*H{row_number})))))'
    )


def _snapshot_cell(cell: Any) -> dict[str, Any]:
    return {
        "style": copy(cell._style),
        "number_format": cell.number_format,
        "alignment": copy(cell.alignment),
        "protection": copy(cell.protection),
    }


def _set_text(cell: Any, value: Any) -> None:
    cell.value = str(value)
    cell.data_type = "s"


def _apply_row_style(worksheet: Any, row_number: int, snapshots: list[dict[str, Any]], height: float | None) -> None:
    for column, snapshot in enumerate(snapshots, start=1):
        cell = worksheet.cell(row=row_number, column=column)
        cell._style = copy(snapshot["style"])
        cell.number_format = snapshot["number_format"]
        cell.alignment = copy(snapshot["alignment"])
        cell.protection = copy(snapshot["protection"])
    worksheet.row_dimensions[row_number].height = height


def _remove_merges_from_row(worksheet: Any, first_row: int) -> None:
    for merged in list(worksheet.merged_cells.ranges):
        if merged.max_row >= first_row:
            worksheet.unmerge_cells(str(merged))


def _add_list_validation(worksheet: Any, cell_range: str, values: list[str]) -> None:
    validation = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=False)
    validation.error = "请选择列表中的有效值。"
    validation.errorTitle = "评分值无效"
    validation.prompt = "请从下拉列表选择。"
    validation.promptTitle = "评分输入"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    worksheet.add_data_validation(validation)
    validation.add(cell_range)


def _export_path(project: Any) -> Path:
    project_id = int(project["id"])
    export_dir = settings.storage_path / "exports" / str(project_id)
    export_dir.mkdir(parents=True, exist_ok=True)
    project_name = _safe_filename_stem(str(project["name"] or "")) or f"project_{project_id}"
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    return export_dir / f"{project_name}_商用密码应用安全性评估打分表_{timestamp}.xlsx"


def _safe_filename_stem(value: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", value).strip(" ._")
    return re.sub(r"_+", "_", cleaned)[:80].strip(" ._")


def _validate_generated_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False, read_only=False)
    expected_names = ["整体测评", "说明", *(item[1] for item in SECTION_LAYOUT)]
    if workbook.sheetnames != expected_names:
        raise ScoreWorkbookExportError("生成的打分表工作表结构不正确。")
    formula_errors: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    if any(marker in value for marker in ("#REF!", "#NAME?", "“", "”")):
                        formula_errors.append(f"{worksheet.title}!{cell.coordinate}")
    if formula_errors:
        raise ScoreWorkbookExportError(f"生成的打分表包含无效公式：{', '.join(formula_errors[:10])}")


def _issue(code: str, unit: str, object_name: str, field: str, message: str) -> dict[str, str]:
    return {
        "section_code": code,
        "unit": unit,
        "object_name": object_name,
        "field": field,
        "message": message,
    }
